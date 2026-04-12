#!/usr/bin/env python3
"""
Parse Hawaii Monthly Visitor Statistics Excel files and extract island-level
data from the "Island" sheet.

Output: hawaii_visitor_stats.csv
Columns: year_month, category, island, value

Usage:
    python parse_visitor_stats.py
"""

import csv
import unicodedata
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import sys
    sys.exit("Missing dependency: pip install openpyxl")

INPUT_DIR  = Path(__file__).parent
OUTPUT_CSV = INPUT_DIR / "hawaii_visitor_stats.csv"

# ── Metric matching (substring of normalized uppercase cell text) ─────────────

METRICS = {
    "TOTAL EXPENDITURES":          "TOTAL EXPENDITURES ($ MILLION)",
    "TOTAL VISITOR DAYS":          "TOTAL VISITOR DAYS",
    "VISITOR ARRIVALS":            "VISITOR ARRIVALS",
    "AVERAGE DAILY CENSUS":        "AVERAGE DAILY CENSUS",
    "AVERAGE LENGTH OF STAY":      "AVERAGE LENGTH OF STAY",
    "PER PERSON PER DAY SPENDING": "PER PERSON PER DAY SPENDING ($)",
    "PER PERSON PER TRIP SPENDING":"PER PERSON PER TRIP SPENDING ($)",
}

# ── Island name normalization ─────────────────────────────────────────────────

def _normalize(s: str) -> str:
    """Strip diacritics and quote-like punctuation, lowercase, collapse whitespace."""
    s = unicodedata.normalize("NFD", s)
    # Remove combining diacritics (Mn) and all apostrophe/quote variants
    s = "".join(
        c for c in s
        if unicodedata.category(c) != "Mn"
        and c not in "'\u2018\u2019\u02bb\u02bc\u0060\u00b4"
    )
    return " ".join(s.lower().split())

ISLAND_MAP = {
    "oahu":          "oahu",
    "maui":          "maui",
    "molokai":       "molokai",
    "lanai":         "lanai",
    "kauai":         "kauai",
    "hawaii island": "hawaii island",
}

def match_island(cell_text: str):
    """Return canonical island name or None if not an island row."""
    n = _normalize(cell_text.strip())
    for key, canonical in ISLAND_MAP.items():
        if key in n:
            return canonical
    return None

def match_metric(cell_text: str):
    """Return canonical metric name or None."""
    upper = cell_text.strip().upper()
    for key, canonical in METRICS.items():
        if key in upper:
            return canonical
    return None

# ── Sheet parser ──────────────────────────────────────────────────────────────

def parse_island_sheet(ws, file_year: int, file_month: int):
    """
    Returns a list of (year_month, category, island, value) tuples.

    For 2025 files: extracts both 2025P and 2024P columns.
    For 2026+ files: extracts only the current-year column.
    """
    rows = list(ws.iter_rows(values_only=True))

    # Find header row: first row where col B looks like "20XXP"
    header_row = None
    for row in rows:
        if row[1] and str(row[1]).strip().endswith("P"):
            header_row = row
            break

    if header_row is None:
        return []

    # Map column index → (year, col_index)
    # Col indices: 0=A, 1=B, 2=C …
    col_years = {}
    import re
    for ci, val in enumerate(header_row):
        if val:
            m = re.match(r"^(\d{4})[A-Z]*P?$", str(val).strip())
            if m:
                col_years[ci] = int(m.group(1))

    # For 2025 files keep 2024P and 2025P; for 2026+ keep only current year
    if file_year == 2025:
        wanted_cols = {ci: yr for ci, yr in col_years.items()}
    else:
        wanted_cols = {ci: yr for ci, yr in col_years.items() if yr == file_year}

    records = []
    current_metric = None

    for row in rows:
        cell_a = str(row[0]).strip() if row[0] is not None else ""
        if not cell_a:
            continue

        # Check if this row is a metric header
        m = match_metric(cell_a)
        if m:
            current_metric = m
            continue

        if current_metric is None:
            continue

        # Check if this is an island row (indented)
        island = match_island(cell_a)
        if island is None:
            continue

        for ci, yr in wanted_cols.items():
            val = row[ci] if ci < len(row) else None
            if val is None:
                continue
            try:
                val = float(val)
            except (TypeError, ValueError):
                continue
            year_month = f"{yr}-{file_month:02d}"
            records.append((year_month, current_metric, island, val))

    return records

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    files = sorted(INPUT_DIR.glob("visitor_stats_*.xlsx"))
    if not files:
        import sys
        sys.exit("No visitor_stats_*.xlsx files found.")

    all_records = []

    for path in files:
        # Parse year and month from filename: visitor_stats_YYYY_MM.xlsx
        parts = path.stem.split("_")
        try:
            file_year  = int(parts[2])
            file_month = int(parts[3])
        except (IndexError, ValueError):
            print(f"  Skipping unrecognised filename: {path.name}")
            continue

        print(f"  Parsing {path.name} ...")
        wb = openpyxl.load_workbook(path, data_only=True)
        if "Island" not in wb.sheetnames:
            print(f"    No 'Island' sheet — skipping.")
            continue

        records = parse_island_sheet(wb["Island"], file_year, file_month)
        print(f"    {len(records)} records extracted.")
        all_records.extend(records)

    # Deduplicate (same year_month + category + island may appear from multiple files)
    seen = {}
    for rec in all_records:
        key = (rec[0], rec[1], rec[2])
        seen[key] = rec  # last file wins

    deduped = sorted(seen.values(), key=lambda r: (r[0], r[1], r[2]))

    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["year_month", "category", "island", "value"])
        writer.writerows(deduped)

    print(f"\nSaved {len(deduped)} rows to {OUTPUT_CSV.name}")


if __name__ == "__main__":
    main()
