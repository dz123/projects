#!/usr/bin/env python3
"""
airbnb_email.py — Sends a Gmail with the YTD Airbnb tax summary (complete
months only) as an HTML table, and attaches the xlsx report.

Usage:
    python airbnb_email.py <path/to/airbnb_taxes_*.xlsx>

Gmail setup:
    Requires a Gmail App Password (Google account → Security → App Passwords).
    Set it via the GMAIL_APP_PASSWORD environment variable, or enter it when
    prompted.
"""

import os
import sys
import getpass
import argparse
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Missing dependency: pip install openpyxl")
    sys.exit(1)

# ── Configuration ─────────────────────────────────────────────────────────────

SENDER        = "danielzhou123@gmail.com"
RECIPIENT     = ["danielzhou123@gmail.com"]
FINAL_EXTRA   = ["daisy.rukawa@gmail.com"]  # added only when run with --final

def get_app_password():
    """Gmail App Password from GMAIL_APP_PASSWORD, else prompted for.

    Never hard-code it here — this repo is public.
    """
    pw = os.environ.get("GMAIL_APP_PASSWORD", "").strip()
    if pw:
        return pw
    try:
        return getpass.getpass(f"Gmail App Password for {SENDER}: ").strip()
    except (EOFError, KeyboardInterrupt):
        return ""

# ── Read summaries from xlsx ──────────────────────────────────────────────────

def read_summaries(xlsx_path: Path):
    """
    Returns a list of dicts, one per complete sheet (sheets with '(incomplete)'
    in the title are skipped), each containing:
        { "heading": "January 2026",
          "headers": ["MV 1321", "PM 2505", "PM 1105", "Total"],
          "rows": [
              ("AIRBNB GROSS (for Tax)", [v1, v2, v3, total]),
              ("Airbnb Fee",             [v1, v2, v3, total]),
              ("Airbnb Net Income",      [v1, v2, v3, total]),
          ]
        }
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    summaries = []

    for ws in wb.worksheets:
        if "(incomplete)" in ws.title.lower():
            continue

        # Row 1: blank | prop1 .. propN | Total  (property count varies)
        headers = []
        c = 2
        while True:
            v = ws.cell(row=1, column=c).value
            if v in (None, ""):
                break
            headers.append(v)
            c += 1
        ncol = len(headers)  # properties + Total

        # Rows 2-4: label | values...
        rows = []
        for r in (2, 3, 4):
            label = ws.cell(row=r, column=1).value
            if not label:
                continue
            values = [ws.cell(row=r, column=cc).value or 0.0
                      for cc in range(2, 2 + ncol)]
            rows.append((label, values))

        if not rows:
            continue

        summaries.append({
            "heading": ws.title,
            "headers": headers,
            "rows":    rows,
        })

    return summaries

# ── HTML generation ───────────────────────────────────────────────────────────

def fmt_money(v):
    if v is None:
        v = 0.0
    sign = "-" if v < 0 else ""
    return f"{sign}${abs(v):,.2f}"

CSS = """
<style>
  body  { font-family: Calibri, Arial, sans-serif; font-size: 13px; color: #222; }
  h2    { margin-top: 28px; margin-bottom: 6px; }
  table { border-collapse: collapse; margin-bottom: 8px; }
  th, td { border: 1px solid #bbb; padding: 5px 12px; text-align: right; }
  th     { background: #2e4057; color: #fff; text-align: center; }
  th.lbl { text-align: left; }
  td.lbl { text-align: left; }
  tr.gross { background: #ffff00; font-weight: bold; }
  tr.net   { background: #e2efda; font-weight: bold; }
  td.note  { border: none; font-style: italic; color: #595959;
             text-align: left; padding-left: 10px; }
</style>
"""

ROW_STYLES = {
    "airbnb gross": "gross",
    "airbnb net income": "net",
}

# Island groupings, matched against the xlsx header labels with spaces removed
# (so both "MV 1321" and "MV1321" work).
ISLANDS = [
    ("Maui", {"MV1321", "KS4113"}),
    ("Oahu", {"PM2505", "PM1105"}),
]

def _row_class(label):
    return ROW_STYLES.get(label.lower().strip(), "")

def _note_cell(label):
    return (" <td class='note'>use this to file taxes</td>"
            if "net income" in label.lower() else "<td class='note'></td>")

def island_groups(headers):
    """[(island, [column indexes])] for the islands present in these headers."""
    groups = []
    for name, members in ISLANDS:
        idx = [i for i, h in enumerate(headers)
               if str(h).replace(" ", "").upper() in members]
        if idx:
            groups.append((name, idx))
    return groups

def build_island_table(s):
    """Per-island table: the two Maui units and the two Oahu units summed."""
    groups = island_groups(s["headers"])
    if not groups:
        return ""

    header_cells = "".join(f"<th>{name}</th>" for name, _ in groups)
    header = (f"<tr><th class='lbl'>By Island: {s['heading']}</th>"
              f"{header_cells}<th>Total</th></tr>")

    data_rows = ""
    for label, values in s["rows"]:
        sums = [sum(values[i] or 0.0 for i in idx) for _, idx in groups]
        cells = "".join(f"<td>{fmt_money(v)}</td>" for v in sums)
        cells += f"<td>{fmt_money(sum(sums))}</td>"
        data_rows += (f"<tr class='{_row_class(label)}'>"
                      f"<td class='lbl'>{label}</td>{cells}{_note_cell(label)}</tr>")

    return f"<table>{header}{data_rows}</table>"

def build_html(summaries):
    sections = []
    for s in summaries:
        heading = s["heading"]
        header_cells = "".join(f"<th>{h}</th>" for h in s["headers"])
        header = (f"<tr><th class='lbl'>Month: {heading}</th>"
                  f"{header_cells}</tr>")

        data_rows = ""
        for label, values in s["rows"]:
            cells = "".join(f"<td>{fmt_money(v)}</td>" for v in values)
            data_rows += (f"<tr class='{_row_class(label)}'>"
                          f"<td class='lbl'>{label}</td>{cells}"
                          f"{_note_cell(label)}</tr>")

        sections.append(
            f"<h2>{heading}</h2>"
            f"<table>{header}{data_rows}</table>"
            f"{build_island_table(s)}"
        )

    return f"<html><head>{CSS}</head><body>{''.join(sections)}</body></html>"

# ── Email sending ─────────────────────────────────────────────────────────────

def send_email(subject, html_body, attachment_path, app_password, recipients):
    msg = MIMEMultipart("mixed")
    msg["From"]    = SENDER
    msg["To"]      = ", ".join(recipients)
    msg["Subject"] = subject

    msg.attach(MIMEText(html_body, "html"))

    with open(attachment_path, "rb") as f:
        part = MIMEBase("application",
                        "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition",
                    f'attachment; filename="{attachment_path.name}"')
    msg.attach(part)

    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        server.ehlo()
        server.starttls()
        server.login(SENDER, app_password)
        server.sendmail(SENDER, recipients, msg.as_string())  # list is fine for sendmail

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Email the latest complete-month Airbnb tax summary.")
    parser.add_argument("xlsx", help="path to airbnb_taxes_*.xlsx")
    parser.add_argument("--final", action="store_true",
                        help=f"also send to {', '.join(FINAL_EXTRA)}")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        sys.exit(f"File not found: {xlsx_path}")

    recipients = RECIPIENT + FINAL_EXTRA if args.final else list(RECIPIENT)

    summaries = read_summaries(xlsx_path)
    if not summaries:
        sys.exit("No complete month sheets found in the xlsx — nothing to email.")

    latest = summaries[-1:]  # only the most recent complete month
    last   = latest[0]["heading"]  # e.g. "Mar 2026"
    subject = f"Airbnb Taxes Report {last}"

    html_body = build_html(latest)

    app_password = get_app_password()
    if not app_password:
        sys.exit("No app password given — set GMAIL_APP_PASSWORD or enter it "
                 "when prompted.")

    print(f"Sending '{subject}' to {', '.join(recipients)} ...")
    send_email(subject, html_body, xlsx_path, app_password, recipients)
    print("Sent.")


if __name__ == "__main__":
    main()
