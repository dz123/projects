#!/usr/bin/env python3
"""
airbnb_report.py — Generates a formatted Excel income report from an Airbnb
transaction CSV, one sheet per month from January through the latest month
found in the file.

Usage:
    python airbnb_report.py [path/to/airbnb_export.csv]

If no path is given, uses the most-recently-modified airbnb_*.csv in Downloads.
Output: Downloads/airbnb_taxes_YYYYMMDD_HHMMSS.xlsx
"""

import sys
import csv
from datetime import datetime, date
from calendar import monthrange
from pathlib import Path

try:
    import pytz
    ET_TZ = pytz.timezone("America/New_York")
except ImportError:
    try:
        from zoneinfo import ZoneInfo
        ET_TZ = ZoneInfo("America/New_York")
    except ImportError:
        ET_TZ = None  # fallback: local time

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing dependency: pip install openpyxl")
    sys.exit(1)

# ── Configuration ────────────────────────────────────────────────────────────

DOWNLOADS  = Path.home() / "Downloads"
OUTPUT_DIR = DOWNLOADS
OWNER      = "Daisy Wu"

LISTING_MAP = {
    "1blk to waikiki beach! newly renovated king condo":    "PM 1105",
    "25thfl ocean view 1blk to waikiki beach king condo":   "PM 2505",
    "partial ocean view king condo at maui vista kihei":    "MV 1321",
    "new ground fl 1bd/2ba @kamaole sands, 3 min beach":     "KS 4113",
}

PROPERTY_FULL = {
    "PM 1105": "Pacific Monarch 1105",
    "PM 2505": "Pacific Monarch 2505",
    "MV 1321": "Maui Vista 1321",
    "KS 4113": "Kamaole Sands 4113",
}

PROPERTIES = ["MV 1321", "PM 2505", "PM 1105", "KS 4113"]

# Confirmation codes to exclude (e.g. displaced bookings replaced by other reservations)
EXCLUDED_CODES = {
    "HMEAQNMMDT",  # Mar 4-17 PM1105 — double-booking, replaced by HM8C54MS8B + HMWYN8JCSK
}

GET_RATE  = 0.04712
TAT_RATE  = 0.11
OTAT_RATE = 0.03

# ── Colours ──────────────────────────────────────────────────────────────────

C_DARK_GREEN  = "1F4E3D"
C_WHITE       = "FFFFFF"
C_LIGHT_GREEN = "E2EFDA"
C_YELLOW      = "FFFF00"
C_CYAN        = "00B0F0"
C_BLACK       = "000000"
C_GRAY        = "595959"

# ── Style helpers ─────────────────────────────────────────────────────────────

MONEY_FMT = '"$"#,##0.00'

def _fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def _style(cell, bold=False, italic=False, size=10,
           bg=None, fg=C_BLACK, h_align="left", fmt=None):
    cell.font      = Font(bold=bold, italic=italic, size=size,
                          color=fg, name="Calibri")
    cell.alignment = Alignment(horizontal=h_align, vertical="center")
    if bg:
        cell.fill = _fill(bg)
    if fmt:
        cell.number_format = fmt

# ── Parsing helpers ───────────────────────────────────────────────────────────

def _date(s):
    try:
        return datetime.strptime(s.strip(), "%m/%d/%Y").date()
    except (ValueError, AttributeError):
        return None

def _float(s):
    try:
        return float((s or "").strip().replace(",", ""))
    except ValueError:
        return 0.0

def _date_range(start, end):
    if end is None:
        return f"{start.strftime('%b')} {start.day}"
    if start.month == end.month and start.year == end.year:
        return f"{start.strftime('%b')} {start.day}-{end.day}"
    return f"{start.strftime('%b')} {start.day} - {end.strftime('%b')} {end.day}"

# ── Data loading ──────────────────────────────────────────────────────────────

def map_listing(listing_str):
    return LISTING_MAP.get(listing_str.strip().lower())

def load_and_process(csv_path: Path):
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))

    # Find latest month from Reservation start dates
    starts = [_date(r["Start date"]) for r in rows
               if r["Type"].strip() == "Reservation" and r.get("Start date", "").strip()]
    starts = [d for d in starts if d]
    if not starts:
        sys.exit("No reservation rows found in CSV.")

    # Error out if any reservation is missing a guest name
    missing = [r["Confirmation code"].strip() for r in rows
               if r["Type"].strip() == "Reservation"
               and not r.get("Guest", "").strip()
               and r["Confirmation code"].strip() not in EXCLUDED_CODES]
    if missing:
        sys.exit(f"Guest name missing for confirmation code(s): {', '.join(missing)}\n"
                 "Re-download the CSV from Airbnb and try again.")

    latest = max(starts)
    year   = latest.year

    # Confirmation code → property (from all reservation rows, all time)
    code_prop: dict[str, str] = {}
    for r in rows:
        if r["Type"].strip() == "Reservation":
            code = r["Confirmation code"].strip()
            prop = map_listing(r.get("Listing", ""))
            if prop and code:
                code_prop[code] = prop

    # Refunds: sum Resolution Adjustment amounts per confirmation code (all time)
    refunds: dict[str, float] = {}
    for r in rows:
        if r["Type"].strip() == "Resolution Adjustment":
            code = r["Confirmation code"].strip()
            if code:
                refunds[code] = refunds.get(code, 0.0) + _float(r.get("Amount", ""))

    # Confirmation codes that have a "Pass Through Tot" row (all time). A genuine
    # income-earning booking always shows up as both a Reservation row and a
    # Pass Through Tot row. A code with a Reservation row but no Pass Through Tot
    # is a cancellation (or a cancelled-and-rebooked duplicate) that earns no
    # income, so we skip those reservations entirely.
    ptt_codes: set[str] = {r["Confirmation code"].strip() for r in rows
                           if r["Type"].strip() == "Pass Through Tot"
                           and r["Confirmation code"].strip()}

    # Build monthly data from January through latest month
    all_data: dict[tuple, dict] = {}

    for month in range(1, latest.month + 1):
        m_start = date(year, month, 1)
        m_end   = date(year, month, monthrange(year, month)[1])
        data: dict[str, list] = {p: [] for p in PROPERTIES}

        for r in rows:
            if r["Type"].strip() != "Reservation":
                continue
            start = _date(r.get("Start date", ""))
            if not start or not (m_start <= start <= m_end):
                continue
            prop = map_listing(r.get("Listing", ""))
            if not prop:
                continue

            code         = r["Confirmation code"].strip()
            if code in EXCLUDED_CODES:
                continue
            if code not in ptt_codes:
                continue  # cancellation: Reservation row with no Pass Through Tot

            amount       = _float(r.get("Amount", ""))
            service_fee  = _float(r.get("Service fee", ""))
            cleaning_fee = _float(r.get("Cleaning fee", ""))
            end          = _date(r.get("End date", ""))

            refund = refunds.get(code, 0.0)

            room_fee     = amount - cleaning_fee + service_fee
            airbnb_fee   = -service_fee
            airbnb_gross = room_fee + cleaning_fee

            data[prop].append({
                "guest":        r["Guest"].strip(),
                "code":         code,
                "start":        start,
                "end":          end,
                "room_fee":     room_fee,
                "cleaning_fee": cleaning_fee,
                "airbnb_gross": airbnb_gross,
                "airbnb_fee":   airbnb_fee,
                "refund":       refund,
            })

        for prop in PROPERTIES:
            data[prop].sort(key=lambda x: x["start"])

        all_data[(year, month)] = data

    return all_data, year, latest.month

# ── Sheet writer ──────────────────────────────────────────────────────────────

TAX_NOTE = "use this to file taxes"

def write_month_sheet(ws, data, m_start, m_end):
    month_label = m_start.strftime("%B %Y").upper()

    # Pre-compute summary per property
    def prop_summary(prop):
        rlist = data[prop]
        gross_for_tax = sum(r["airbnb_gross"] + r["refund"] for r in rlist)
        airbnb_fee    = sum(r["airbnb_fee"]                 for r in rlist)
        net_income    = gross_for_tax + airbnb_fee
        get_t  = sum(r["airbnb_gross"] * GET_RATE  for r in rlist)
        tat_t  = sum(r["airbnb_gross"] * TAT_RATE  for r in rlist)
        otat_t = sum(r["airbnb_gross"] * OTAT_RATE for r in rlist)
        return dict(gross_for_tax=gross_for_tax, airbnb_fee=airbnb_fee,
                    net_income=net_income, get=get_t, tat=tat_t, otat=otat_t)

    summary = {p: prop_summary(p) for p in PROPERTIES}
    tot = {k: sum(summary[p][k] for p in PROPERTIES)
           for k in ("gross_for_tax", "airbnb_fee", "net_income", "get", "tat", "otat")}

    # Column of the summary "Total" (after the per-property columns) and the
    # note that follows it — derived so they track the number of properties.
    total_col = 2 + len(PROPERTIES)
    note_col  = total_col + 1

    max_guests = max((len(data[p]) for p in PROPERTIES), default=0)
    ws.column_dimensions["A"].width = 24
    for ci in range(2, 2 + max_guests + 3):
        ws.column_dimensions[get_column_letter(ci)].width = 15

    cur = 1

    # ── TOP SUMMARY ────────────────────────────────────────────────────────────

    for col, label in enumerate(["", *PROPERTIES, "Total"], 1):
        _style(ws.cell(row=cur, column=col, value=label), bold=True, h_align="center")
    cur += 1

    # AIRBNB GROSS (for Tax) — note in column 6 (first empty after Total)
    _style(ws.cell(row=cur, column=1, value="AIRBNB GROSS  (for Tax)"),
           bold=True, bg=C_YELLOW)
    for i, prop in enumerate(PROPERTIES):
        _style(ws.cell(row=cur, column=2+i, value=summary[prop]["gross_for_tax"]),
               h_align="right", fmt=MONEY_FMT, bg=C_YELLOW)
    _style(ws.cell(row=cur, column=total_col, value=tot["gross_for_tax"]),
           bold=True, h_align="right", fmt=MONEY_FMT, bg=C_YELLOW)
    cur += 1

    # Airbnb Fee
    _style(ws.cell(row=cur, column=1, value="Airbnb Fee"))
    for i, prop in enumerate(PROPERTIES):
        _style(ws.cell(row=cur, column=2+i, value=summary[prop]["airbnb_fee"]),
               h_align="right", fmt=MONEY_FMT)
    _style(ws.cell(row=cur, column=total_col, value=tot["airbnb_fee"]),
           bold=True, h_align="right", fmt=MONEY_FMT)
    cur += 1

    # Airbnb Net Income
    _style(ws.cell(row=cur, column=1, value="Airbnb Net Income"),
           bold=True, bg=C_LIGHT_GREEN)
    for i, prop in enumerate(PROPERTIES):
        _style(ws.cell(row=cur, column=2+i, value=summary[prop]["net_income"]),
               h_align="right", fmt=MONEY_FMT, bg=C_LIGHT_GREEN)
    _style(ws.cell(row=cur, column=total_col, value=tot["net_income"]),
           bold=True, h_align="right", fmt=MONEY_FMT, bg=C_LIGHT_GREEN)
    _style(ws.cell(row=cur, column=note_col, value=TAX_NOTE),
           italic=True, fg=C_GRAY, bg=C_LIGHT_GREEN)
    cur += 1

    # Tax type labels
    cur += 1
    for col, label in enumerate(["GET", "TAT", "OTAT"], 2):
        _style(ws.cell(row=cur, column=col, value=label),
               italic=True, fg=C_GRAY, h_align="center")
    cur += 1

    # Total Tax values
    _style(ws.cell(row=cur, column=1, value="Total Tax"), italic=True)
    for col, val in enumerate([tot["get"], tot["tat"], tot["otat"]], 2):
        _style(ws.cell(row=cur, column=col, value=val),
               h_align="right", fmt=MONEY_FMT)
    cur += 3

    # ── PER-PROPERTY DETAIL ────────────────────────────────────────────────────

    for prop in PROPERTIES:
        rlist = data[prop]
        n     = len(rlist)
        tcol  = 2 + n   # total column index

        for label in [f"Unit:  {PROPERTY_FULL[prop]}",
                      f"Owner:  {OWNER}",
                      f"Month:  {month_label}"]:
            _style(ws.cell(row=cur, column=4, value=label), bold=True, h_align="center")
            cur += 1
        cur += 1

        _style(ws.cell(row=cur, column=1, value="Reservations"), bold=True)
        cur += 1

        if n == 0:
            ws.cell(row=cur, column=1, value="(no reservations this month)")
            cur += 4
            continue

        # Guest name headers
        _style(ws.cell(row=cur, column=1, value="Name of Guest"))
        for i, res in enumerate(rlist):
            _style(ws.cell(row=cur, column=2+i, value=res["guest"]),
                   bold=True, bg=C_DARK_GREEN, fg=C_WHITE, h_align="center")
        cur += 1

        # Date Stayed
        _style(ws.cell(row=cur, column=1, value="Date Stayed"))
        for i, res in enumerate(rlist):
            _style(ws.cell(row=cur, column=2+i,
                           value=_date_range(res["start"], res["end"])),
                   h_align="center")
        cur += 1

        def money_row(label, vals, total=None, bold_lbl=False, bg=None, note=None):
            nonlocal cur
            _style(ws.cell(row=cur, column=1, value=label), bold=bold_lbl, bg=bg)
            for i, v in enumerate(vals):
                _style(ws.cell(row=cur, column=2+i, value=v),
                       h_align="right", fmt=MONEY_FMT, bg=bg)
            if total is not None:
                _style(ws.cell(row=cur, column=tcol, value=total),
                       bold=True, h_align="right", fmt=MONEY_FMT, bg=bg)
            if note:
                _style(ws.cell(row=cur, column=tcol+1, value=note),
                       italic=True, fg=C_GRAY)
            cur += 1

        room_fees   = [r["room_fee"]     for r in rlist]
        clean_fees  = [r["cleaning_fee"] for r in rlist]
        gross_vals  = [r["airbnb_gross"] for r in rlist]
        fee_vals    = [r["airbnb_fee"]   for r in rlist]
        refund_vals = [r["refund"]       for r in rlist]
        net_vals    = [r["airbnb_gross"] + r["airbnb_fee"] + r["refund"]
                       for r in rlist]

        money_row("Room Fee",     room_fees,  sum(room_fees))
        money_row("Cleaning Fee", clean_fees, sum(clean_fees))
        money_row("AIRBNB GROSS", gross_vals, sum(gross_vals),
                  bold_lbl=True, bg=C_YELLOW)
        cur += 1

        money_row("Airbnb Fee", fee_vals, sum(fee_vals))
        if any(v != 0 for v in refund_vals):
            money_row("Refund to Guest", refund_vals, sum(refund_vals))
        money_row("Airbnb Net Income", net_vals, sum(net_vals),
                  bold_lbl=True, bg=C_LIGHT_GREEN, note=TAX_NOTE)
        cur += 1

        _style(ws.cell(row=cur, column=1, value="Taxes"), bold=True)
        cur += 1

        gets  = [r["airbnb_gross"] * GET_RATE  for r in rlist]
        tats  = [r["airbnb_gross"] * TAT_RATE  for r in rlist]
        otats = [r["airbnb_gross"] * OTAT_RATE for r in rlist]
        ttax  = [gets[i] + tats[i] + otats[i]  for i in range(n)]

        money_row(f"GET ({GET_RATE*100:.3f}%)",   gets,  sum(gets))
        money_row(f"TAT ({TAT_RATE*100:.0f}%)",   tats,  sum(tats))
        money_row(f"OTAT ({OTAT_RATE*100:.0f}%)", otats, sum(otats))

        for i, v in enumerate(ttax):
            _style(ws.cell(row=cur, column=2+i, value=v),
                   bold=True, h_align="right", fmt=MONEY_FMT)
        _style(ws.cell(row=cur, column=tcol, value=sum(ttax)),
               bold=True, h_align="right", fmt=MONEY_FMT, bg=C_CYAN, fg=C_WHITE)
        cur += 2

        payouts = [net_vals[i] + ttax[i] for i in range(n)]
        money_row("Total Host Payout", payouts, sum(payouts),
                  bold_lbl=True, bg=C_LIGHT_GREEN)

        cur += 3

# ── Excel file generation ─────────────────────────────────────────────────────

def generate_excel(all_data, year, latest_month):
    if ET_TZ:
        now_et = datetime.now(ET_TZ)
    else:
        now_et = datetime.now()

    out_name = f"airbnb_taxes_{now_et.strftime('%Y%m%d_%H%M%S')}.xlsx"
    out_path = OUTPUT_DIR / out_name

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove the default blank sheet

    today = date.today()
    for month in range(1, latest_month + 1):
        m_start = date(year, month, 1)
        m_end   = date(year, month, monthrange(year, month)[1])
        title   = m_start.strftime("%b %Y")
        if month == latest_month and today < m_end:
            title += " (incomplete)"
        ws = wb.create_sheet(title=title)
        write_month_sheet(ws, all_data[(year, month)], m_start, m_end)

    wb.active = wb.worksheets[-1]
    wb.save(out_path)
    return out_path

# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) > 1:
        csv_path = Path(sys.argv[1])
    else:
        # Match only timestamped exports: airbnb_YYYYMMDD_HHMMSS.csv
        d8, d6 = "[0-9]" * 8, "[0-9]" * 6
        candidates = sorted(DOWNLOADS.glob(f"airbnb_{d8}_{d6}.csv"),
                            key=lambda p: p.stat().st_mtime, reverse=True)
        if not candidates:
            sys.exit("No airbnb_YYYYMMDD_HHMMSS.csv found in Downloads. "
                     "Pass the CSV file path as an argument.")
        csv_path = candidates[0]
        print(f"Using: {csv_path.name}")

    if not csv_path.exists():
        sys.exit(f"File not found: {csv_path}")

    all_data, year, latest_month = load_and_process(csv_path)
    out_path = generate_excel(all_data, year, latest_month)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
